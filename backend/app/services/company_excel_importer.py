from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO
import json
import re
import uuid
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "公司基本信息",
    "软著",
    "专利已授权",
    "专利审核中",
    "专业资质认证",
    "管理体系",
    "双软证书",
    "荣誉证书",
    "其他认证证书",
    "人员",
    "业绩",
]

CERTIFICATE_SHEETS = {
    "专业资质认证",
    "管理体系",
    "双软证书",
    "荣誉证书",
    "其他认证证书",
}

ASSET_TYPES = {
    "软著": "software_copyright",
    "专利已授权": "patent_granted",
    "专利审核中": "patent_pending",
    "人员": "personnel_certificate",
    "业绩": "project_case",
}


class CompanyExcelImporter:
    """Parse the company qualification workbook into normalized evidence items."""

    def parse(self, file_bytes: bytes) -> Dict[str, Any]:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        batch_id = str(uuid.uuid4())
        warnings: List[str] = []
        assets: List[Dict[str, Any]] = []

        missing_sheets = [sheet for sheet in EXPECTED_SHEETS if sheet not in wb.sheetnames]
        for sheet in missing_sheets:
            warnings.append(f"缺少 Sheet: {sheet}")

        company_name = self._parse_company_name(wb) or ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name == "公司基本信息":
                continue
            if sheet_name in CERTIFICATE_SHEETS:
                assets.extend(self._parse_certificate_sheet(ws, sheet_name, company_name, batch_id, warnings))
            elif sheet_name == "软著":
                assets.extend(self._parse_software_sheet(ws, company_name, batch_id, warnings))
            elif sheet_name in {"专利已授权", "专利审核中"}:
                assets.extend(self._parse_patent_sheet(ws, sheet_name, company_name, batch_id, warnings))
            elif sheet_name == "人员":
                assets.extend(self._parse_personnel_sheet(ws, company_name, batch_id, warnings))
            elif sheet_name == "业绩":
                assets.extend(self._parse_project_sheet(ws, company_name, batch_id, warnings))

        warnings.extend(self._duplicate_warnings(assets))
        summary = self._build_summary(assets, warnings)
        return {
            "batch_id": batch_id,
            "company_name": company_name,
            "assets": assets,
            "summary": summary,
            "warnings": warnings,
        }

    def _parse_company_name(self, wb) -> Optional[str]:
        if "公司基本信息" not in wb.sheetnames:
            return None
        ws = wb["公司基本信息"]
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            values = [self._clean(v) for v in row]
            for index, value in enumerate(values):
                if value == "公司名称" and index + 1 < len(values):
                    return values[index + 1]
        return None

    def _parse_certificate_sheet(self, ws, sheet_name: str, company_name: str, batch_id: str, warnings: List[str]) -> List[Dict[str, Any]]:
        rows = self._rows_as_dicts(ws)
        assets = []
        for row_no, row in rows:
            name = self._clean(row.get("资质名称"))
            if not name:
                continue
            asset = self._base_asset(
                company_name=company_name or self._clean(row.get("公司名称")),
                asset_type="qualification",
                source_sheet=sheet_name,
                name=name,
                batch_id=batch_id,
                row=row,
            )
            asset.update(
                category=self._clean(row.get("资质分类")) or sheet_name,
                certificate_no=self._clean(row.get("证书编号")),
                issuer=self._clean(row.get("发证机构")),
                issue_date=self._date(row.get("发证日期")),
                expiry_date=self._date(row.get("有效期至")),
                status=self._status(row.get("当前状态"), row.get("有效期至")),
            )
            self._warn_invalid_date(sheet_name, row_no, "有效期至", row.get("有效期至"), asset["expiry_date"], warnings)
            assets.append(asset)
        return assets

    def _parse_software_sheet(self, ws, company_name: str, batch_id: str, warnings: List[str]) -> List[Dict[str, Any]]:
        rows = self._rows_as_dicts(ws)
        assets = []
        for row_no, row in rows:
            name = self._clean(row.get("软件全称"))
            if not name:
                continue
            asset = self._base_asset(company_name or self._clean(row.get("公司名称")), "software_copyright", "软著", name, batch_id, row)
            asset.update(
                certificate_no=self._clean(row.get("登记号")) or self._clean(row.get("证书号")),
                issuer=self._clean(row.get("发证机构")),
                issue_date=self._date(row.get("发证日期")),
                expiry_date=self._date(row.get("有效期至")),
                status=self._status(row.get("是否有效"), row.get("有效期至")),
            )
            self._warn_invalid_date("软著", row_no, "有效期至", row.get("有效期至"), asset["expiry_date"], warnings)
            assets.append(asset)
        return assets

    def _parse_patent_sheet(self, ws, sheet_name: str, company_name: str, batch_id: str, warnings: List[str]) -> List[Dict[str, Any]]:
        rows = self._rows_as_dicts(ws)
        assets = []
        for row_no, row in rows:
            name = self._clean(row.get("专利名称"))
            if not name:
                continue
            asset = self._base_asset(company_name or self._clean(row.get("公司名称")), ASSET_TYPES[sheet_name], sheet_name, name, batch_id, row)
            asset.update(
                category=self._clean(row.get("专利类型")),
                certificate_no=self._clean(row.get("专利号")) or self._clean(row.get("证书编号")),
                issuer=self._clean(row.get("发证机构")),
                issue_date=self._date(row.get("获证日期")) or self._date(row.get("专利申请时间")),
                expiry_date=self._date(row.get("有效期至")),
                status=self._status(row.get("是否有效"), row.get("有效期至")) if sheet_name == "专利已授权" else "审核中",
            )
            self._warn_invalid_date(sheet_name, row_no, "有效期至", row.get("有效期至"), asset["expiry_date"], warnings)
            assets.append(asset)
        return assets

    def _parse_personnel_sheet(self, ws, company_name: str, batch_id: str, warnings: List[str]) -> List[Dict[str, Any]]:
        rows = self._rows_as_dicts(ws)
        assets = []
        for row_no, row in rows:
            name = self._clean(row.get("资质证书"))
            person = self._clean(row.get("姓名"))
            if not name and not person:
                continue
            asset = self._base_asset(company_name, "personnel_certificate", "人员", name, batch_id, row)
            asset.update(
                category=self._clean(row.get("类别")),
                certificate_no=self._clean(row.get("资质证书编号")),
                issuer=self._clean(row.get("颁发机构")),
                issue_date=self._date(row.get("批准日期")),
                expiry_date=self._date(row.get("到期时间")),
                status=self._status(None, row.get("到期时间")),
            )
            self._warn_invalid_date("人员", row_no, "到期时间", row.get("到期时间"), asset["expiry_date"], warnings)
            assets.append(asset)
        return assets

    def _parse_project_sheet(self, ws, company_name: str, batch_id: str, warnings: List[str]) -> List[Dict[str, Any]]:
        rows = self._rows_as_dicts(ws)
        assets = []
        for _, row in rows:
            name = self._clean(row.get("合同名称"))
            if not name:
                continue
            asset = self._base_asset(company_name, "project_case", "业绩", name, batch_id, row)
            asset.update(
                category=self._clean(row.get("客户类型")),
                issue_date=self._date(row.get("签订时间")),
                expiry_date=self._date(row.get("终止时间")),
                status="有效",
                amount_wanyuan=self._amount_wanyuan(row.get("合同金额")),
            )
            assets.append(asset)
        return assets

    def _rows_as_dicts(self, ws) -> List[Tuple[int, Dict[str, Any]]]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [self._clean(value) for value in rows[0]]
        result = []
        for row_no, values in enumerate(rows[1:], start=2):
            row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
            if any(value not in (None, "") for value in row.values()):
                result.append((row_no, row))
        return result

    def _base_asset(self, company_name: str, asset_type: str, source_sheet: str, name: str, batch_id: str, row: Dict[str, Any]) -> Dict[str, Any]:
        data = {key: self._jsonable(value) for key, value in row.items()}
        keywords = self._keywords(" ".join(str(v) for v in data.values() if v not in (None, "")))
        return {
            "company_name": company_name or "",
            "asset_type": asset_type,
            "source_sheet": source_sheet,
            "name": name,
            "category": None,
            "certificate_no": None,
            "issuer": None,
            "issue_date": None,
            "expiry_date": None,
            "status": None,
            "amount_wanyuan": None,
            "keywords": keywords,
            "data": data,
            "import_batch_id": batch_id,
        }

    def _build_summary(self, assets: List[Dict[str, Any]], warnings: List[str]) -> Dict[str, Any]:
        by_type = Counter(asset["asset_type"] for asset in assets)
        by_sheet = Counter(asset["source_sheet"] for asset in assets)
        expired = [asset for asset in assets if asset.get("status") == "过期"]
        expiring_soon = [
            asset for asset in assets
            if asset.get("expiry_date") and asset.get("status") == "有效" and asset["expiry_date"] <= datetime.now() + timedelta(days=180)
        ]
        valid_qualifications = [
            asset["name"] for asset in assets
            if asset["asset_type"] == "qualification" and asset.get("status") == "有效"
        ]
        return {
            "total_assets": len(assets),
            "by_type": dict(by_type),
            "by_sheet": dict(by_sheet),
            "valid_qualification_count": len(valid_qualifications),
            "expired_count": len(expired),
            "expiring_soon_count": len(expiring_soon),
            "warning_count": len(warnings),
            "top_qualifications": valid_qualifications[:20],
        }

    def _duplicate_warnings(self, assets: Iterable[Dict[str, Any]]) -> List[str]:
        keys = [
            asset.get("certificate_no") or f"{asset.get('source_sheet')}::{asset.get('name')}"
            for asset in assets
            if asset.get("certificate_no") or asset.get("name")
        ]
        counts = Counter(keys)
        return [f"重复资料: {key} 出现 {count} 次" for key, count in counts.items() if count > 1]

    def _warn_invalid_date(self, sheet: str, row_no: int, field: str, raw: Any, parsed: Optional[datetime], warnings: List[str]) -> None:
        if raw in (None, "", "/", "-", "—", "长期", "永久"):
            return
        if parsed is None:
            warnings.append(f"{sheet} 第 {row_no} 行 {field} 无法解析: {raw}")

    def _status(self, raw_status: Any, raw_expiry: Any) -> str:
        status = self._clean(raw_status)
        expiry = self._date(raw_expiry)
        if "过期" in status:
            return "过期"
        if "审核" in status:
            return "审核中"
        if expiry and expiry < datetime.now():
            return "过期"
        if status in {"有效", "正常"} or expiry or raw_expiry in {"长期", "永久", "—", "-", "/"}:
            return "有效"
        return status or "未知"

    def _date(self, value: Any) -> Optional[datetime]:
        if isinstance(value, datetime):
            return value
        if isinstance(value, (int, float)):
            try:
                return datetime(1899, 12, 30) + timedelta(days=float(value))
            except Exception:
                return None
        text = self._clean(value)
        if not text or text in {"/", "-", "—", "长期", "永久", "未发表"}:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%Y-%m", "%Y/%m", "%Y.%m"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    def _amount_wanyuan(self, value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            amount = float(str(value).replace(",", ""))
        except ValueError:
            return None
        return round(amount / 10000, 2)

    def _keywords(self, text: str) -> str:
        tokens = set()
        keyword_map = {
            "AI/人工智能": ["AI", "人工智能", "大模型", "深度学习", "智能"],
            "大数据": ["大数据", "数据治理", "数据分析", "数据资产", "数据平台"],
            "软件开发": ["软件", "系统", "平台", "开发", "SaaS"],
            "安全/信创": ["安全", "信创", "麒麟", "统信", "鲲鹏", "海光", "等保"],
            "运维/服务": ["运维", "维护", "技术服务", "服务"],
            "通信/网络": ["通信", "5G", "网络", "移动", "电信"],
        }
        upper_text = text.upper()
        for token, aliases in keyword_map.items():
            if any(alias.upper() in upper_text for alias in aliases):
                tokens.add(token)
        return json.dumps(sorted(tokens), ensure_ascii=False)

    def _jsonable(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return value

    def _clean(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        return re.sub(r"\s+", " ", text)
